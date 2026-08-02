#!/usr/bin/env python3
"""
The synchronization layer — the ONLY module besides workbook.py that opens
master.xlsx. Every other script in this project reads plain JSON that this
module produces; none of them import openpyxl or know a spreadsheet exists.

Two directions:
  read  — master.xlsx Zone-1 sheets -> data/sync/*.json (what scripts consume)
  write — fresh market data -> the hidden _MarketCache sheet (what Dashboard
          formulas read). Zone-1 sheets are NEVER written by this direction —
          only a human (or Claude, acting on the human's instruction during a
          session) edits Portfolio/Transactions/Watchlist/etc.

Four more directions exist for agents that need to RECORD, CORRECT, or
REORGANIZE something during a session without every agent needing its own
openpyxl knowledge — these are the ONLY way anything outside this file
writes to the workbook; the openpyxl logic stays here, callable, not
duplicated:
  append — one JSON row (or, with --rows, a JSON array of rows in one
           file save — for bulk imports) -> appended to a named Zone-1 sheet.
  update — set columns on every row matching --match (correcting a wrong
           value, resolving a Notes-sheet question) without deleting history.
  delete — remove every row matching --match (a genuinely wrong row, e.g. a
           migration-time inference that turned out incorrect — for a
           RESOLVED QUESTION rather than a mistake, prefer `update` to set
           status=resolved instead, so the resolution stays visible).
  sort   — re-order a sheet's rows by one column's value (e.g. group
           Portfolio by instrument_type). Pure re-ordering, nothing added
           or removed; Portfolio's type color-coding is conditional
           formatting (see workbook.py) so it doesn't need re-applying.

`write-cache` also auto-appends today's computed total portfolio value to
the hidden _ValueHistory sheet (and data/valuations.csv) every time it
runs — see sync.py's `_log_value_history()`.

Usage:
  python data/sync/sync.py read                  # xlsx -> JSON
  python data/sync/sync.py write-cache            # fresh snapshot -> _MarketCache + _ValueHistory
  python data/sync/sync.py append --sheet Watchlist --row '{"ticker": "V", ...}'
  python data/sync/sync.py update --sheet Notes --match '{"id": "3"}' --row '{"status": "resolved", "resolution": "..."}'
  python data/sync/sync.py delete --sheet Portfolio --match '{"account_id": "swedbank-fund"}'
  python data/sync/sync.py sort --sheet Portfolio --by instrument_type
  python data/sync/sync.py read --xlsx other.xlsx --out-dir /tmp/x
"""
import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from schema import SHEETS, ZONE1_SHEETS  # noqa: E402

from openpyxl import load_workbook

DEFAULT_XLSX = "master.xlsx"
DEFAULT_OUT_DIR = "data/sync"
UNIVERSE_CACHE_PATH = "data/cache/universe.json"  # machine-owned; sp500 auto-built here
THESIS_CACHE_PATH = "data/cache/thesis_candidates.json"  # regenerated from Watchlist+Thesis each sync

# sheet name -> output JSON filename (snake_case, matches what scripts expect)
SHEET_TO_FILE = {
    "Portfolio": "portfolio.json",
    "Transactions": "transactions.json",
    "Watchlist": "watchlist.json",
    "Universe": "universe_manual.json",
    "Investment Thesis": "thesis.json",
    "Pending Orders": "pending_orders.json",
    "Settings": "settings.json",
    "Notes": "notes.json",
    "Manual Data": "manual_data.json",
}


def _sheet_rows(ws, columns):
    """Yield dict rows from row 2 onward; skip fully-blank rows."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or v == "" for v in row):
            continue
        rec = {}
        for i, col in enumerate(columns):
            val = row[i] if i < len(row) else None
            rec[col] = val
        yield rec


def read_workbook(xlsx_path, out_dir):
    """Zone 1: master.xlsx -> JSON files every other module consumes."""
    if not os.path.exists(xlsx_path):
        sys.exit(f"{xlsx_path} not found. Run data/sync/workbook.py first, "
                 f"then scripts/migrate_from_json.py to populate it.")
    wb = load_workbook(xlsx_path, data_only=False)
    os.makedirs(out_dir, exist_ok=True)

    written = {}
    for sheet_name in ZONE1_SHEETS:
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Expected sheet '{sheet_name}' missing from {xlsx_path} — "
                     f"workbook is out of date, rebuild with data/sync/workbook.py")
        rows = list(_sheet_rows(wb[sheet_name], SHEETS[sheet_name]))
        out_file = os.path.join(out_dir, SHEET_TO_FILE[sheet_name])
        with open(out_file, "w") as f:
            json.dump({
                "synced_utc": datetime.now(timezone.utc).isoformat(),
                "source_sheet": sheet_name,
                "row_count": len(rows),
                "rows": rows,
            }, f, indent=2, default=str)
        written[sheet_name] = (out_file, len(rows))

    print(f"Synced {xlsx_path} -> {out_dir}/")
    for sheet, (path, n) in written.items():
        print(f"  {sheet:<20s} {n:>4d} rows -> {path}")

    _merge_universe_manual(written.get("Universe"))
    _rebuild_thesis_cache(written.get("Watchlist"), written.get("Investment Thesis"))
    return written


def _rebuild_thesis_cache(watchlist_written, thesis_written):
    """rank_candidates.py reads a candidates cache in the same shape the old
    thesis_candidates.json used — regenerate it from the synced Watchlist
    (which tickers are being tracked) + Investment Thesis (the prose/risk_tag
    for each) sheets every sync, so it never drifts from what's in Excel."""
    if not watchlist_written or not thesis_written:
        return
    with open(watchlist_written[0]) as f:
        watchlist = json.load(f)["rows"]
    with open(thesis_written[0]) as f:
        thesis_by_ticker = {t["ticker"]: t for t in json.load(f)["rows"] if t.get("ticker")}

    candidates = []
    for w in watchlist:
        t = thesis_by_ticker.get(w.get("ticker"), {})
        candidates.append({
            "ticker": w.get("ticker"), "name": w.get("name"), "source": w.get("source"),
            "date": w.get("date_added"), "risk_tag": t.get("risk_tag"),
            "thesis": t.get("thesis"), "policy_tailwind": t.get("policy_tailwind"),
        })
    out = {"note": "Regenerated from the Watchlist + Investment Thesis sheets each sync — not hand-edited.",
           "candidates": candidates}
    os.makedirs(os.path.dirname(THESIS_CACHE_PATH), exist_ok=True)
    with open(THESIS_CACHE_PATH, "w") as f:
        json.dump(out, f, indent=2)
    print(f"  rebuilt thesis cache -> {THESIS_CACHE_PATH} ({len(candidates)} candidates)")


def _merge_universe_manual(universe_written):
    """The Universe SHEET is the human-owned source for manually-curated
    tickers (Nordic, Europe, thesis picks); data/cache/universe.json is
    machine-owned (build_universe.py writes its auto-built sp500 list there).
    Rather than have both independently hold manual categories — the exact
    'same information in two places' the architecture forbids — this merges
    the sheet's rows into the cache file's manual categories on every sync,
    leaving sp500 untouched. rank_candidates.py / screen_candidates.py then
    read ONE file and don't need to know Excel exists."""
    if not universe_written:
        return
    path, _n = universe_written
    with open(path) as f:
        manual = json.load(f)["rows"]

    if os.path.exists(UNIVERSE_CACHE_PATH):
        with open(UNIVERSE_CACHE_PATH) as f:
            cache = json.load(f)
    else:
        cache = {"categories": {}, "metadata": {}}

    cats, meta = cache.setdefault("categories", {}), cache.setdefault("metadata", {})
    manual_cat_names = sorted({r["category"] for r in manual if r.get("category")})
    for cat in manual_cat_names:
        cats[cat] = sorted({r["ticker"] for r in manual if r.get("category") == cat and r.get("ticker")})
    for r in manual:
        if r.get("ticker"):
            meta[r["ticker"]] = {"name": r.get("name"), "sector": meta.get(r["ticker"], {}).get("sector"),
                                 "cik": meta.get(r["ticker"], {}).get("cik"),
                                 "currency": r.get("currency"), "source": r.get("source")}
    cache["manual_categories"] = manual_cat_names
    os.makedirs(os.path.dirname(UNIVERSE_CACHE_PATH), exist_ok=True)
    with open(UNIVERSE_CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)
    print(f"  merged {len(manual)} manual universe rows -> {UNIVERSE_CACHE_PATH} "
          f"(categories: {manual_cat_names}; sp500 untouched)")


def _log_value_history(wb, portfolio_rows, records):
    """Append one row to the hidden _ValueHistory sheet (and a mirrored row
    to data/valuations.csv for scripts/performance.py) with TODAY's total
    portfolio value. This is the "keeps getting updated automatically" piece
    — previously valuations.csv only grew when a human/agent remembered to
    append a row by hand.

    TBD (unlisted-fund) rows are a special case: they have no `quantity`
    (there's no live per-unit price to multiply it by), so their
    _MarketCache record carries `market_value_sek` as the row's TOTAL value
    (= cost_basis_total_sek) directly, not a per-unit price. Every other row
    type uses quantity x market_value_sek. Mixing these up (i.e. always
    multiplying by quantity) silently zeroes out every unlisted fund — a
    real bug caught 2026-08-02 when this function's first version did
    exactly that and undercounted the portfolio by ~130k SEK."""
    pf_cols = SHEETS["Portfolio"]
    ticker_idx = pf_cols.index("ticker")
    name_idx = pf_cols.index("name")
    qty_idx = pf_cols.index("quantity")
    total = 0.0
    for row in portfolio_rows:
        t = row[ticker_idx]
        if t == "TBD":
            key = f"TBD-{row[name_idx]}"
            rec = records.get(key) or {}
            mv = rec.get("market_value_sek")
            if isinstance(mv, (int, float)):
                total += mv
            continue
        rec = records.get(t) or {}
        qty, mv = row[qty_idx], rec.get("market_value_sek")
        if isinstance(qty, (int, float)) and isinstance(mv, (int, float)):
            total += qty * mv

    vh = wb["_ValueHistory"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    r = vh.max_row + 1
    vh.cell(row=r, column=1, value=today)
    vh.cell(row=r, column=2, value=round(total, 2))
    vh.cell(row=r, column=3, value="auto-logged by run.py fetch")

    valuations_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "data", "valuations.csv")
    is_new = not os.path.exists(valuations_path)
    with open(valuations_path, "a", newline="") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(["date", "total_value_sek", "net_contribution_since_last_sek", "note"])
        w.writerow([today, round(total, 2), 0, "auto-logged by run.py fetch"])
    return round(total, 2)


def write_market_cache(xlsx_path, records):
    """Zone 2: fresh {ticker: {last_price, currency, price_as_of, market_value_sek,
    fetch_status, data_source}} -> the hidden _MarketCache sheet. Overwrites the
    sheet wholesale every call — it is fully disposable, exactly like the JSON
    cache under data/cache/. Never touches any Zone-1 sheet (except appending
    to _ValueHistory, itself Zone 2)."""
    wb = load_workbook(xlsx_path)
    ws = wb["_MarketCache"]
    ws.delete_rows(2, ws.max_row)  # clear all data rows, keep the header
    cols = SHEETS["_MarketCache"]
    for i, (ticker, rec) in enumerate(records.items(), 2):
        for j, col in enumerate(cols, 1):
            val = ticker if col == "ticker" else rec.get(col)
            ws.cell(row=i, column=j, value=val)

    portfolio_rows = [[c.value for c in row] for row in wb["Portfolio"].iter_rows(min_row=2)
                      if any(c.value not in (None, "") for c in row)]
    total = _log_value_history(wb, portfolio_rows, records)
    print(f"Logged total value {total} SEK to _ValueHistory and data/valuations.csv")

    dash = wb["Dashboard"]
    dash["B17"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb.save(xlsx_path)
    print(f"Wrote {len(records)} rows to _MarketCache in {xlsx_path}")


def append_rows(xlsx_path, sheet_name, rows):
    """Append MANY rows (list of dicts, keys matching schema.py's column
    list — unknown keys ignored, missing keys left blank) to a Zone-1 sheet
    in ONE file open/save cycle. The only way an agent or bulk-import script
    records new information into master.xlsx; keeps the openpyxl/file-format
    knowledge in this one module. Caller should follow up with `read` if the
    appended data needs to flow to the other JSON caches (e.g. a new
    Watchlist row needs `read` to rebuild the thesis cache).

    Built for bulk imports — e.g. flattened values copied out of Excel's
    Stocks data type into Manual Data, dozens of ticker/field rows at once —
    where doing one load/save per row would be needlessly slow and risks a
    half-written file if interrupted partway."""
    if sheet_name not in ZONE1_SHEETS:
        sys.exit(f"'{sheet_name}' is not a Zone-1 sheet — append only writes "
                 f"human-owned input sheets. Zone-1 sheets: {ZONE1_SHEETS}")
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    cols = SHEETS[sheet_name]
    next_row = ws.max_row + 1
    for row in rows:
        for i, col in enumerate(cols, 1):
            ws.cell(row=next_row, column=i, value=row.get(col))
        next_row += 1
    wb.save(xlsx_path)
    print(f"Appended {len(rows)} row(s) to '{sheet_name}' in {xlsx_path}")
    print("Run 'python data/sync/sync.py read' to flow this into data/sync/*.json "
          "and any dependent caches (universe/thesis).")


def append_row(xlsx_path, sheet_name, row):
    """Append exactly one row — thin wrapper over append_rows() for the
    common single-row case (existing CLI/agent usage unchanged)."""
    append_rows(xlsx_path, sheet_name, [row])


def _matching_rows(ws, cols, match):
    """Yield (row_number, record_dict) for every data row (from row 2) whose
    columns match every key in `match`."""
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        rec = dict(zip(cols, values))
        if all(rec.get(k) == v for k, v in match.items()):
            yield row[0].row, rec


def update_rows(xlsx_path, sheet_name, match, updates):
    """Update every row in a Zone-1 sheet whose columns match `match` (dict)
    by setting the columns in `updates` (dict). append-only was a real gap —
    Zone-1 sheets sometimes need CORRECTING (a migration-time inference that
    turned out wrong, a stale note) not just growing, and the only prior
    alternative was opening master.xlsx directly, which breaks 'only sync.py
    understands Excel'. Returns the count of rows updated."""
    if sheet_name not in ZONE1_SHEETS:
        sys.exit(f"'{sheet_name}' is not a Zone-1 sheet — update only writes "
                 f"human-owned input sheets. Zone-1 sheets: {ZONE1_SHEETS}")
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    cols = SHEETS[sheet_name]
    col_index = {c: i + 1 for i, c in enumerate(cols)}
    updated = 0
    for row_num, _rec in list(_matching_rows(ws, cols, match)):
        for k, v in updates.items():
            if k in col_index:
                ws.cell(row=row_num, column=col_index[k], value=v)
        updated += 1
    wb.save(xlsx_path)
    print(f"Updated {updated} row(s) in '{sheet_name}' matching {match} in {xlsx_path}")
    if updated == 0:
        print("  WARNING: no rows matched — nothing changed. Check --match against the sheet's actual values.")
    print("Run 'python data/sync/sync.py read' to flow this into data/sync/*.json.")
    return updated


def delete_rows(xlsx_path, sheet_name, match):
    """Delete every row in a Zone-1 sheet whose columns match `match`. Use
    for genuinely wrong rows (a migration-time inference error, a
    duplicate) — for a RESOLVED QUESTION rather than a mistake, prefer
    `update` to set status=resolved (Notes-sheet convention) over deleting,
    so the resolution history stays visible. Returns the count deleted."""
    if sheet_name not in ZONE1_SHEETS:
        sys.exit(f"'{sheet_name}' is not a Zone-1 sheet — delete only writes "
                 f"human-owned input sheets. Zone-1 sheets: {ZONE1_SHEETS}")
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    cols = SHEETS[sheet_name]
    to_delete = [row_num for row_num, _rec in _matching_rows(ws, cols, match)]
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r, 1)
    wb.save(xlsx_path)
    print(f"Deleted {len(to_delete)} row(s) from '{sheet_name}' matching {match} in {xlsx_path}")
    if not to_delete:
        print("  WARNING: no rows matched — nothing changed.")
    print("Run 'python data/sync/sync.py read' to flow this into data/sync/*.json.")
    return len(to_delete)


def sort_sheet(xlsx_path, sheet_name, by):
    """Re-order a Zone-1 sheet's data rows by one column's value (e.g. group
    Portfolio by instrument_type: stocks together, funds together, etc.).
    Pure re-ordering — no data changed, no rows added/removed. Formatting
    (conditional-formatting fills, header style) already targets the whole
    range so it doesn't need re-applying after a sort. Rows with equal `by`
    values keep their prior relative order (stable sort)."""
    if sheet_name not in ZONE1_SHEETS:
        sys.exit(f"'{sheet_name}' is not a Zone-1 sheet — sort only reorders "
                 f"human-owned input sheets. Zone-1 sheets: {ZONE1_SHEETS}")
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    cols = SHEETS[sheet_name]
    if by not in cols:
        sys.exit(f"'{by}' is not a column of '{sheet_name}'. Columns: {cols}")
    by_idx = cols.index(by)
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)
            if any(c.value not in (None, "") for c in row)]
    rows.sort(key=lambda r: (r[by_idx] is None, r[by_idx]))
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    wb.save(xlsx_path)
    print(f"Sorted {len(rows)} row(s) in '{sheet_name}' by '{by}' in {xlsx_path}")
    return len(rows)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("direction", choices=["read", "write-cache", "append", "update", "delete", "sort"])
    p.add_argument("--xlsx", default=DEFAULT_XLSX)
    p.add_argument("--out-dir", default=DEFAULT_OUT_DIR)
    p.add_argument("--sheet", help="(append/update/delete/sort) target sheet name")
    p.add_argument("--row", help="(append) JSON object for the new row; (update) JSON object of columns to set")
    p.add_argument("--rows", help="(append) JSON array of row objects — bulk append in one file save")
    p.add_argument("--match", help="(update/delete) JSON object of column:value a row must match")
    p.add_argument("--by", help="(sort) column name to group/sort the sheet's rows by")
    args = p.parse_args()

    if args.direction == "read":
        read_workbook(args.xlsx, args.out_dir)
    elif args.direction == "append":
        if not args.sheet or not (args.row or args.rows):
            sys.exit("append requires --sheet and either --row '<json object>' or --rows '<json array>'")
        if args.rows:
            append_rows(args.xlsx, args.sheet, json.loads(args.rows))
        else:
            append_row(args.xlsx, args.sheet, json.loads(args.row))
    elif args.direction == "update":
        if not args.sheet or not args.match or not args.row:
            sys.exit("update requires --sheet, --match '<json object>', and --row '<json object of columns to set>'")
        update_rows(args.xlsx, args.sheet, json.loads(args.match), json.loads(args.row))
    elif args.direction == "delete":
        if not args.sheet or not args.match:
            sys.exit("delete requires --sheet and --match '<json object>'")
        delete_rows(args.xlsx, args.sheet, json.loads(args.match))
    elif args.direction == "sort":
        if not args.sheet or not args.by:
            sys.exit("sort requires --sheet and --by <column name>")
        sort_sheet(args.xlsx, args.sheet, args.by)
    else:
        cache_path = os.path.join(args.out_dir, "market_cache.json")
        if not os.path.exists(cache_path):
            sys.exit(f"{cache_path} not found — run.py should write it before "
                     f"calling 'write-cache' (see run.py's fetch step).")
        with open(cache_path) as f:
            records = json.load(f)
        write_market_cache(args.xlsx, records)


if __name__ == "__main__":
    main()
