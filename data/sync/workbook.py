#!/usr/bin/env python3
"""
Builds a fresh master.xlsx with all sheets, headers, and the Dashboard's
formulas. Run standalone to create an EMPTY workbook (headers only) — the
normal path is to run scripts/migrate_from_json.py afterward to populate it
from the current JSON state.

This is the ONLY module besides sync.py that constructs sheet layout — schema.py
is the shared column definition both import from.

Usage:
  python data/sync/workbook.py                    # writes master.xlsx (headers only)
  python data/sync/workbook.py --out other.xlsx    # write elsewhere (for tests)
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from schema import SHEETS, ZONE1_SHEETS, ZONE2_SHEETS  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import LineChart, Reference

HEADER_FILL = PatternFill(start_color="1F6E63", end_color="1F6E63", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DASHBOARD_LABEL_FONT = Font(bold=True, size=12)
DASHBOARD_SECTION_FONT = Font(bold=True, size=14, color="1F6E63")

# instrument_type -> row fill, applied to the Portfolio sheet via conditional
# formatting keyed on the instrument_type column (F) — self-maintaining: a
# new row gets colored correctly regardless of where it's added or whether
# the sheet has been re-sorted, no script re-run needed to stay correct.
INSTRUMENT_TYPE_FILLS = {
    "stock": "DCEBF7",        # light blue
    "fund": "E2F0D9",         # light green
    "certificate": "FCE4D6",  # light orange
    "cash": "F2F2F2",         # light gray
    "spot_crypto": "E6DCF7",  # light purple
}


def _apply_portfolio_type_formatting(ws, ncols):
    """Color each Portfolio row by its instrument_type (column F), so
    stocks/funds/certificates/cash/crypto are visually distinct at a glance
    without needing the sheet re-sorted or a script re-run every time a row
    is added — conditional formatting evaluates live, per row."""
    last_col_letter = get_column_letter(ncols)
    for itype, color in INSTRUMENT_TYPE_FILLS.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            f"A2:{last_col_letter}1000",
            FormulaRule(formula=[f'$F2="{itype}"'], fill=fill),
        )


def _style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def ensure_zone2_sheet(wb, name):
    """Create a Zone-2 (hidden, machine-write) sheet with headers if it
    doesn't already exist. Idempotent — safe to call on a workbook that may
    already have it, which is what lets a retrofit script add a new Zone-2
    sheet to an already-populated master.xlsx without touching Zone-1 data."""
    if name in wb.sheetnames:
        return wb[name]
    cols = SHEETS[name]
    ws = wb.create_sheet(name)
    for i, col in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=col)
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)
    _style_header_row(ws, len(cols))
    ws.sheet_state = "hidden"
    return ws


def build_dashboard(wb):
    """(Re)build the Dashboard sheet in an already-open Workbook. Pure view,
    formula-only — safe to delete and recreate at any time since nothing here
    is human-owned data. Assumes Portfolio, _MarketCache, and _ValueHistory
    already exist in `wb`. Used both by build_workbook() (fresh workbook) and
    by retrofit scripts (refresh the Dashboard on an already-populated file)."""
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)  # index 0 = first tab
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 32
    dash.column_dimensions["B"].width = 18
    dash.column_dimensions["C"].width = 40

    dash["A1"] = "Finance Council — Dashboard"
    dash["A1"].font = Font(bold=True, size=18, color="1F6E63")
    dash["A2"] = "Computed from Portfolio + _MarketCache. Never edit this sheet by hand."
    dash["A2"].font = Font(italic=True, size=9, color="5C6669")

    # TBD (unlisted-fund) rows have no live per-unit price to multiply a
    # quantity by — their _MarketCache record carries market_value_sek as
    # the row's TOTAL (= cost_basis_total_sek) directly, set by run.py's
    # cmd_fetch. Every formula below that sums "quantity x price" must
    # therefore EXCLUDE ticker="TBD" from that product and separately ADD
    # cost_basis_total_sek (column L) for those rows via SUMIF/SUMIFS —
    # otherwise unlisted funds (a real ~130k SEK of this portfolio) silently
    # vanish from every total, a bug caught and fixed 2026-08-02.
    PRICED = '(Portfolio!A2:A1000<>"TBD")*Portfolio!J2:J1000*IFERROR(VLOOKUP(Portfolio!A2:A1000,_MarketCache!A:E,5,0),0)'

    dash["A4"] = "PORTFOLIO TOTAL"
    dash["A4"].font = DASHBOARD_SECTION_FONT
    dash["A5"] = "Total value (SEK)"
    dash["B5"] = f'=SUMPRODUCT({PRICED})+SUMIF(Portfolio!A2:A1000,"TBD",Portfolio!L2:L1000)'
    dash["A5"].font = DASHBOARD_LABEL_FONT
    dash["C5"] = ("Sum of each holding's quantity x current price (from _MarketCache), "
                  "plus cost basis for unlisted funds (never get a live price)")
    dash["C5"].font = Font(italic=True, size=9, color="5C6669")

    dash["A7"] = "RISK TIER BREAKDOWN"
    dash["A7"].font = DASHBOARD_SECTION_FONT
    tiers = [("Secure", "secure", 60), ("Medium", "medium", 30), ("High-risk", "high", 10)]
    row = 8
    dash.cell(row=row, column=1, value="Tier")
    dash.cell(row=row, column=2, value="Value (SEK)")
    dash.cell(row=row, column=3, value="Target %")
    for c in range(1, 4):
        dash.cell(row=row, column=c).font = Font(bold=True)
    for label, key, target in tiers:
        row += 1
        dash.cell(row=row, column=1, value=label)
        dash.cell(row=row, column=2,
                  value=(f'=(SUMPRODUCT((Portfolio!H2:H1000="{key}")*{PRICED})'
                         f'+SUMIFS(Portfolio!L2:L1000,Portfolio!H2:H1000,"{key}",Portfolio!A2:A1000,"TBD"))'
                         f'/$B$5'))
        dash.cell(row=row, column=3, value=f"{target}%")

    dash["A13"] = "FEE DRAG"
    dash["A13"].font = DASHBOARD_SECTION_FONT
    dash["A14"] = "Total annual fees (SEK/yr)"
    dash["A14"].font = DASHBOARD_LABEL_FONT
    dash["B14"] = (f'=SUMPRODUCT((Portfolio!A2:A1000<>"TBD")*Portfolio!M2:M1000/100*'
                   f'Portfolio!J2:J1000*IFERROR(VLOOKUP(Portfolio!A2:A1000,_MarketCache!A:E,5,0),0))'
                   f'+SUMPRODUCT((Portfolio!A2:A1000="TBD")*Portfolio!M2:M1000/100*Portfolio!L2:L1000)')
    dash["C14"] = "Sum of annual_fee_pct x market value per holding (cost basis for unlisted funds)"
    dash["C14"].font = Font(italic=True, size=9, color="5C6669")

    dash["A16"] = "DATA FRESHNESS"
    dash["A16"].font = DASHBOARD_SECTION_FONT
    dash["A17"] = "Last sync"
    dash["A17"].font = DASHBOARD_LABEL_FONT
    dash["B17"] = "(set by run.py sync)"

    dash["A19"] = "BY TYPE"
    dash["A19"].font = DASHBOARD_SECTION_FONT
    row = 20
    dash.cell(row=row, column=1, value="Type")
    dash.cell(row=row, column=2, value="Value (SEK)")
    dash.cell(row=row, column=3, value="Count")
    for c in range(1, 4):
        dash.cell(row=row, column=c).font = Font(bold=True)
    for label, key in [("Stocks", "stock"), ("Funds", "fund"),
                        ("Certificates", "certificate"), ("Cash", "cash"),
                        ("Crypto", "spot_crypto")]:
        row += 1
        dash.cell(row=row, column=1, value=label)
        dash.cell(row=row, column=2,
                  value=(f'=SUMPRODUCT((Portfolio!F2:F1000="{key}")*{PRICED})'
                         f'+SUMIFS(Portfolio!L2:L1000,Portfolio!F2:F1000,"{key}",Portfolio!A2:A1000,"TBD")'))
        dash.cell(row=row, column=3, value=f'=COUNTIF(Portfolio!F2:F1000,"{key}")')
        dash.cell(row=row, column=1).fill = PatternFill(
            start_color=INSTRUMENT_TYPE_FILLS[key], end_color=INSTRUMENT_TYPE_FILLS[key],
            fill_type="solid")

    perf_start = row + 2
    dash.cell(row=perf_start, column=1, value="PERFORMANCE OVER TIME")
    dash.cell(row=perf_start, column=1).font = DASHBOARD_SECTION_FONT
    dash.cell(row=perf_start + 1, column=1,
              value="See chart below — logged automatically by `run.py fetch` into "
                    "the hidden _ValueHistory sheet each time it runs.")
    dash.cell(row=perf_start + 1, column=1).font = Font(italic=True, size=9, color="5C6669")

    chart = LineChart()
    chart.title = "Total portfolio value (SEK)"
    chart.y_axis.title = "SEK"
    chart.x_axis.title = "Date"
    chart.width, chart.height = 22, 9
    vh = wb["_ValueHistory"]
    data_ref = Reference(vh, min_col=2, min_row=1, max_row=200)  # total_value_sek incl. header
    dates_ref = Reference(vh, min_col=1, min_row=2, max_row=200)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(dates_ref)
    dash.add_chart(chart, f"A{perf_start + 3}")

    footer_row = perf_start + 22
    dash.cell(row=footer_row, column=1,
              value="Add new KPIs below this line — this sheet is designed to grow.")
    dash.cell(row=footer_row, column=1).font = Font(italic=True, size=9, color="8FA09D")
    return dash


def build_workbook(out_path):
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet, we add our own in order

    # --- Zone 1: human-owned input sheets ---
    for name in ["Portfolio", "Transactions", "Watchlist", "Universe",
                 "Investment Thesis", "Pending Orders", "Settings", "Notes",
                 "Manual Data"]:
        cols = SHEETS[name]
        ws = wb.create_sheet(name)
        for i, col in enumerate(cols, 1):
            ws.cell(row=1, column=i, value=col)
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)
        _style_header_row(ws, len(cols))
        if name == "Portfolio":
            _apply_portfolio_type_formatting(ws, len(cols))

    # --- Zone 2: machine-write cache, hidden ---
    for name in ZONE2_SHEETS:
        ensure_zone2_sheet(wb, name)

    # --- Dashboard: pure view, formulas only, reads Portfolio + _MarketCache ---
    build_dashboard(wb)

    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  Zone 1 (human-owned): {', '.join(ZONE1_SHEETS)}")
    print(f"  Zone 2 (machine cache, hidden): {', '.join(ZONE2_SHEETS)}")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="master.xlsx")
    args = p.parse_args()
    build_workbook(args.out)
