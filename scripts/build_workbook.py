#!/usr/bin/env python3
"""
Generate master.xlsx from portfolio.json + the latest snapshot.

WHY THIS EXISTS, AND WHY IT ONLY WRITES
---------------------------------------
The Excel-backed branch made master.xlsx the *source of truth*: the user
edited the workbook and a sync layer read it back. That inverts the burden
onto the person, and the user's explicit requirement is the opposite -
"I want to minimize the strain on me, I do not want a lot to update and feed."

So the direction is reversed here. portfolio.json stays the source of truth
(agents maintain it), and this script renders it into a workbook you only
ever *look at*. Nothing reads Holdings/Overview back. You can delete
master.xlsx at any time and regenerate it; you can never break the system
by editing those sheets.

FORMULA-BASED DASHBOARD (2026-08-04)
-------------------------------------
Overview used to be Python-computed static numbers. It's now a thin raw-data
sheet (Holdings) plus real Excel formulas (SUMIF/SUMPRODUCT) on Overview that
compute totals, allocation, and drift FROM that raw data - inspectable,
auditable, and correct by construction as long as Holdings' facts are right.
Python still owns every FACT (price, quantity, value - anything that traces
to a fetch or a user statement); Excel owns every COMBINATION of facts
(totals, percentages, drift). Same "never invent a number" boundary as
before, just drawn between two sheets instead of hidden inside a script.

TWO SHEETS SURVIVE A REBUILD, VERBATIM
----------------------------------------
"Manual Data" and "Fundamentals" are never regenerated - if they already
exist, their exact current contents (values, not formulas) are carried
forward, and only genuinely new tickers get an appended blank row. This is
what lets you safely put hand-entered or Excel-Stocks-data-type-sourced
figures directly in the workbook without a rebuild wiping them out.

THE "FUNDAMENTALS" TAB AND EXCEL'S STOCKS DATA TYPE
-----------------------------------------------------
Non-US fundamentals (P/E, sector, market cap for Nordic tickers) have no
free automated source this system can reach - confirmed empirically, see
OPEN_ITEMS.md's closed items. But Excel's own "Stocks" data type (a
Microsoft 365 feature, LSEG-sourced) can pull exactly this, live, in YOUR
Excel session - it just can't be triggered or read by anything headless.
The Fundamentals tab is the bridge: type/convert tickers to the Stocks data
type in your own Excel, pull whatever fields you want into the row for that
ticker, save. Next time this script runs, it reads back whatever plain
values are sitting in those cells (openpyxl cannot read a live Stocks link -
only the last value Excel itself resolved and cached) and carries them
forward as static values. Run `scripts/import_fundamentals_tab.py` to fold
those values into data/company_profiles/<TICKER>.json, the cache the
swedish-equity-review skill actually reads.

LIMITATION, STATED PLAINLY: this is a periodic manual refresh, not a live
feed. Every time this script (or any tool) re-saves master.xlsx, any STILL-
LIVE Stocks formula in it is at real risk of losing its live link (openpyxl
does not preserve Rich/Linked Data Types - see SYSTEM.md's investigation
note). If you want fresh numbers, redo the Stocks lookup in your own Excel
and save again; don't expect it to silently keep updating itself here.

Usage:
    python scripts/build_workbook.py
    python scripts/build_workbook.py --out master.xlsx
"""

import argparse
import glob
import json
import os
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SNAP_DIR = os.path.join(ROOT, "data", "cache", "snapshots")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4EC")
OK_FILL = PatternFill("solid", fgColor="E8F5E9")

# Wide-format columns for the user-maintained Fundamentals tab - matches how
# Excel's Stocks data type naturally spills fields (one row per ticker, one
# column per field), not the long ticker/field/value shape Manual Data used
# in the archived Excel branch. Extend this list freely; unknown extra
# columns a user adds by hand are preserved too, this is just the seed set.
FUNDAMENTALS_COLS = ["ticker", "name", "price", "pe_ratio", "market_cap", "sector",
                     "industry", "dividend_yield_pct", "week52_high", "week52_low",
                     "beta", "as_of", "notes"]


def load_json(path):
    with open(path) as fh:
        return json.load(fh)


def latest_snapshot():
    files = sorted(glob.glob(os.path.join(SNAP_DIR, "*.json")))
    if not files:
        return None, None
    return os.path.basename(files[-1]), load_json(files[-1])


def write_header(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, max_width=52):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def fx_rate(snap, key):
    """SEK per unit of foreign currency, from the snapshot's macro block only."""
    node = ((snap or {}).get("macro") or {}).get(key) or {}
    try:
        return float(node.get("value"))
    except (TypeError, ValueError):
        return None


def holding_value(h, snap):
    """Live value where possible, else the recorded value. Never invents one."""
    eq = (snap or {}).get("equities", {})
    t = h.get("ticker")
    qty = h.get("quantity")

    # Cash is held as a quantity in its own currency; convert non-SEK at the
    # snapshot's rate. Without a fetched rate we report no data rather than
    # guessing - a wrong FX rate silently distorts every allocation figure.
    if h.get("instrument_type") == "cash" and qty is not None:
        if t == "CASH_SEK" or h.get("currency") == "SEK":
            return qty, "book"
        rate = fx_rate(snap, "sek_per_usd" if h.get("currency") == "USD" else "sek_per_eur")
        return (qty * rate, "live FX") if rate else (None, "no data")

    if t in eq and not eq[t].get("error") and eq[t].get("price") and qty is not None:
        return eq[t]["price"] * qty, "live"
    if h.get("market_value_sek") is not None:
        return h["market_value_sek"], "user-relayed" if h.get("market_value_source") else "book"
    return None, "no data"


def build_holdings(wb, pf, snap, rows):
    """RAW DATA sheet - one row per holding, stable column layout, nothing
    but fetched/computed/user-relayed facts. Overview's formulas reference
    this sheet by fixed column letter; nothing here is itself a formula."""
    ws = wb.create_sheet("Holdings")
    cols = ["Ticker", "Name", "Account", "Exposure class", "Instrument type",
            "Qty", "Value (SEK)", "Cost basis (SEK)", "Fee %/yr", "Price source"]
    write_header(ws, cols)
    last_row = 1
    for i, r in enumerate(rows, start=2):
        last_row = i
        ws.cell(row=i, column=1, value=r["ticker"])
        ws.cell(row=i, column=2, value=r["name"])
        ws.cell(row=i, column=3, value=r["account"])
        ws.cell(row=i, column=4, value=r["cls"])
        ws.cell(row=i, column=5, value=r["itype"])
        ws.cell(row=i, column=6, value=r["qty"])
        vc = ws.cell(row=i, column=7, value=r["value"])
        vc.number_format = "#,##0.00"
        cc = ws.cell(row=i, column=8, value=r["cost"])
        if isinstance(r["cost"], (int, float)):
            cc.number_format = "#,##0.00"
        ws.cell(row=i, column=9, value=r["fee"])
        sc = ws.cell(row=i, column=10, value=r["source"])
        if r["source"] == "no data":
            sc.fill = BAD_FILL
        elif r["source"] != "live":
            sc.fill = WARN_FILL
        # Gain % as a real formula (value vs cost), not Python arithmetic -
        # IFERROR covers "no data"/missing-cost/zero-cost cleanly.
        gc = ws.cell(row=i, column=11, value=f"=IFERROR((G{i}-H{i})/H{i},\"no data\")")
        gc.number_format = "+0.0%;-0.0%"
    ws.cell(row=1, column=11, value="Δ vs cost").fill = HEADER_FILL
    ws.cell(row=1, column=11).font = HEADER_FONT
    autosize(ws)
    return last_row


def build_overview(wb, pf, last_holdings_row):
    """Formula-only aggregation over Holdings. Python's only job here is
    writing the TARGET percentages (from portfolio.json, a fact) - every
    total, share, and drift figure is a live Excel formula referencing
    Holdings, so changing a raw fact recalculates everything downstream
    without rerunning this script."""
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "Portfolio overview"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} from portfolio.json"
    ws["A3"] = "Every number below is a live formula over the Holdings sheet - edit portfolio.json to change a fact, not this file."
    ws["A3"].font = Font(italic=True, size=9)
    ws["A4"] = "READ-ONLY VIEW — edit portfolio.json, not this file. Rebuild: python scripts/build_workbook.py"
    ws["A4"].font = Font(italic=True)

    hr = last_holdings_row  # last data row on Holdings, for range formulas
    VAL = f"Holdings!$G$2:$G${hr}"
    CLS = f"Holdings!$D$2:$D${hr}"
    # Total counted value excludes exposure_class "excluded_operating_cash"
    # (Revolut etc - real money, not investable capital, see portfolio.json).
    TOTAL = f'SUMIF({CLS},"<>excluded_operating_cash",{VAL})'

    ws["A6"] = "Allocation vs target"
    ws["A6"].font = Font(bold=True)
    write_header(ws, ["Exposure class", "Value (SEK)", "% of total", "Target %", "Drift (pp)"], row=7)

    targets = pf.get("targets", {})
    tmap = {"equity": targets.get("equity_pct"), "crypto": targets.get("crypto_pct"),
            "cash": targets.get("cash_pct"), "fixed_income": targets.get("fixed_income_pct")}
    classes = sorted({h.get("exposure_class") for h in pf.get("holdings", [])
                      if h.get("exposure_class") and h.get("exposure_class") != "excluded_operating_cash"})

    total_row = 7 + len(classes) + 1  # header(7) + one row per class + TOTAL, fixed in advance
    r = 7
    for cls in classes:
        r += 1
        ws.cell(row=r, column=1, value=cls)
        vc = ws.cell(row=r, column=2, value=f'=SUMIF({CLS},"{cls}",{VAL})')
        vc.number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=f'=IFERROR(B{r}/$B${total_row}*100,"n/a")').number_format = "0.0"
        tgt = tmap.get(cls)
        ws.cell(row=r, column=4, value=tgt if tgt is not None else "n/a")
        if tgt is not None:
            dc = ws.cell(row=r, column=5, value=f'=IFERROR(C{r}-D{r},"n/a")')
            dc.number_format = "+0.0;-0.0"
            # Conditional-formatting-free color: can't evaluate the formula
            # result at write time, so this uses openpyxl conditional
            # formatting instead of a static fill (added below, once).
        else:
            ws.cell(row=r, column=5, value="n/a")
    assert r + 1 == total_row, "class-row count drifted from the pre-computed total_row"
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    tc = ws.cell(row=total_row, column=2, value=f"={TOTAL}")
    tc.font = Font(bold=True)
    tc.number_format = "#,##0.00"

    # Color drift by magnitude - conditional formatting so it re-evaluates
    # if the user opens this in real Excel and something upstream changes.
    from openpyxl.formatting.rule import CellIsRule
    drift_range = f"E8:E{total_row - 1}"
    ws.conditional_formatting.add(drift_range, CellIsRule(operator="greaterThanOrEqual", formula=["10"], fill=BAD_FILL))
    ws.conditional_formatting.add(drift_range, CellIsRule(operator="lessThanOrEqual", formula=["-10"], fill=BAD_FILL))
    ws.conditional_formatting.add(drift_range, CellIsRule(operator="between", formula=["3", "9.99"], fill=WARN_FILL))
    ws.conditional_formatting.add(drift_range, CellIsRule(operator="between", formula=["-9.99", "-3"], fill=WARN_FILL))
    ws.conditional_formatting.add(drift_range, CellIsRule(operator="between", formula=["-2.99", "2.99"], fill=OK_FILL))

    r = total_row + 2
    ws.cell(row=r, column=1, value="BY TYPE").font = Font(bold=True)
    r += 1
    write_header(ws, ["Instrument type", "Value (SEK)", "Count"], row=r)
    ITYPE = f"Holdings!$E$2:$E${hr}"
    types = sorted({h.get("instrument_type") for h in pf.get("holdings", [])
                    if h.get("exposure_class") != "excluded_operating_cash" and h.get("instrument_type")})
    for itype in types:
        r += 1
        ws.cell(row=r, column=1, value=itype)
        vc = ws.cell(row=r, column=2, value=f'=SUMIF({ITYPE},"{itype}",{VAL})')
        vc.number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=f'=COUNTIF({ITYPE},"{itype}")')

    r += 2
    ws.cell(row=r, column=1, value="FEE DRAG").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Total annual fees (SEK/yr)").font = Font(bold=True, size=12)
    FEE = f"Holdings!$I$2:$I${hr}"
    fc = ws.cell(row=r, column=2, value=f'=SUMPRODUCT(({FEE}<>"")*IFERROR({FEE}/100,0)*IFERROR({VAL},0))')
    fc.number_format = "#,##0.00"
    ws.cell(row=r, column=3, value="Sum of fee%/yr x value per holding, over Holdings")
    ws.cell(row=r, column=3).font = Font(italic=True, size=9, color="5C6669")

    r += 2
    ws.cell(row=r, column=1, value="CROSS-CHECK").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="If TOTAL above does not match your broker, the system's model is wrong — "
                                   "that mismatch is the point of this sheet.")
    r += 1
    excluded = [h.get("name") for h in pf.get("holdings", [])
                if h.get("exposure_class") == "excluded_operating_cash"]
    if excluded:
        ws.cell(row=r, column=1, value="Deliberately excluded (operating cash, not capital): " + "; ".join(excluded))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    return ws


def build_history(wb):
    path = os.path.join(ROOT, "data", "valuations.csv")
    if not os.path.exists(path):
        return
    import csv
    ws = wb.create_sheet("Value history")
    with open(path) as fh:
        for i, row in enumerate(csv.reader(fh), start=1):
            for j, v in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=v)
                if i == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
    ws.column_dimensions["D"].width = 60
    for c in "ABC":
        ws.column_dimensions[c].width = 18
    autosize(ws, max_width=60)


def _read_existing_sheet_values(xlsx_path, sheet_name):
    """Best-effort read of a sheet's cell VALUES (data_only=True: cached
    results, never a formula string) from an already-saved workbook. Used to
    carry Manual Data / Fundamentals forward across a rebuild. Returns None
    if the file/sheet doesn't exist or can't be read (corrupt/locked) - a
    rebuild must never be blocked by this."""
    if not os.path.exists(xlsx_path):
        return None
    try:
        old = load_workbook(xlsx_path, data_only=True)
        if sheet_name not in old.sheetnames:
            return None
        return [[c.value for c in row] for row in old[sheet_name].iter_rows()]
    except Exception as exc:
        print(f"note: could not read existing '{sheet_name}' sheet ({exc}) — rebuilding fresh")
        return None


def build_manual_data(wb, preserved):
    ws = wb.create_sheet("Manual Data")
    if preserved:
        for row in preserved:
            ws.append(row)
        print(f"preserved {len(preserved) - 1} manual data row(s)")
    else:
        write_header(ws, ["ticker", "field", "value", "currency", "as_of", "source", "notes"])
        ws.cell(row=2, column=7, value="Hand-entered figures go here; this sheet survives rebuilds.")
    autosize(ws)


def build_fundamentals(wb, pf, preserved):
    """User-maintained tab for figures Excel's Stocks data type can reach
    but this pipeline can't (non-US fundamentals). Preserved verbatim across
    rebuilds; new holdings get an appended blank row (ticker+name only) so
    there's always something to fill in for a position that doesn't have one
    yet. Never overwrites a row that already exists for a ticker."""
    ws = wb.create_sheet("Fundamentals")
    existing_tickers = set()
    if preserved:
        for row in preserved:
            ws.append(row)
            if len(row) > 0 and row[0] and row[0] != FUNDAMENTALS_COLS[0]:
                existing_tickers.add(row[0])
        print(f"preserved {len(preserved) - 1} fundamentals row(s)")
    else:
        write_header(ws, FUNDAMENTALS_COLS)
        ws.cell(row=2, column=len(FUNDAMENTALS_COLS),
                value="Use Excel's Stocks data type on the ticker, pull whatever fields you want into "
                      "this row's columns, save. This sheet survives rebuilds. Run "
                      "scripts/import_fundamentals_tab.py to fold values into data/company_profiles/.")

    held_tickers = {(h.get("ticker"), h.get("name")) for h in pf.get("holdings", [])
                    if h.get("ticker") and h.get("instrument_type") in ("stock", "certificate")
                    and h.get("ticker") not in ("TBD",)}
    next_row = ws.max_row + 1
    added = 0
    for ticker, name in sorted(held_tickers):
        if ticker in existing_tickers:
            continue
        ws.cell(row=next_row, column=1, value=ticker)
        ws.cell(row=next_row, column=2, value=name)
        next_row += 1
        added += 1
    if added:
        print(f"added {added} new ticker row(s) to Fundamentals (blank, ready to fill)")
    autosize(ws)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(ROOT, "master.xlsx"))
    args = ap.parse_args()

    pf = load_json(os.path.join(ROOT, "data", "portfolio.json"))
    snap_name, snap = latest_snapshot()

    rows = []
    for h in pf.get("holdings", []):
        if h.get("quantity") == 0:
            continue
        val, source = holding_value(h, snap)
        cost = h.get("cost_basis_total_sek")
        if cost is None and h.get("cost_basis_per_unit") is not None and h.get("quantity") is not None:
            cost = h["cost_basis_per_unit"] * h["quantity"]
        rows.append({
            "ticker": h.get("ticker"),
            "name": h.get("name"),
            "account": h.get("account_id"),
            "cls": h.get("exposure_class"),
            "itype": h.get("instrument_type"),
            "qty": h.get("quantity") if h.get("quantity") is not None else "n/a",
            "value": val if val is not None else "no data",
            "cost": cost if cost is not None else "no data",
            "fee": h.get("annual_fee_pct") if h.get("annual_fee_pct") is not None else "n/a",
            "source": source,
        })

    preserved_manual = _read_existing_sheet_values(args.out, "Manual Data")
    preserved_fundamentals = _read_existing_sheet_values(args.out, "Fundamentals")

    wb = Workbook()
    wb.remove(wb.active)
    last_row = build_holdings(wb, pf, snap, rows)
    build_overview(wb, pf, last_row)
    build_history(wb)
    build_manual_data(wb, preserved_manual)
    build_fundamentals(wb, pf, preserved_fundamentals)

    wb.save(args.out)
    counted = [r for r in rows if isinstance(r["value"], (int, float)) and r["cls"] != "excluded_operating_cash"]
    total = sum(r["value"] for r in counted)
    print(f"Wrote {args.out}  ({len(counted)} positions counted, total {total:,.2f} SEK)")


if __name__ == "__main__":
    main()
