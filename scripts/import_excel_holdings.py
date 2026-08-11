#!/usr/bin/env python3
"""
Fold a user-maintained Excel workbook (Holdings/Transactions/FX Rates/
Watchlist, built on Excel's live "Stocks" data type) into this system's
actual state files: data/company_profiles/<TICKER>.json, data/portfolio.json
holdings, data/transactions.csv, and data/cache/watchlist.json.

WHY THIS EXISTS
----------------
Non-US fundamentals (P/E, sector, market cap, employee count, etc. for
Nordic tickers) have no free automated source this system can reach headless
- confirmed empirically, see OPEN_ITEMS.md's closed items. Excel's own
"Stocks" data type (Microsoft 365, LSEG-sourced) can pull exactly this,
live, in the user's own Excel session. This script is the bridge: read the
workbook's CACHED values (data_only=True - this cannot see or preserve a
live Stocks link, only whatever Excel itself last resolved and cached) and
fold them into the system's real state.

READ-ONLY, DELIBERATELY. This script never opens the source workbook in
write mode and never saves back to it. Two real, still-unexplained quirks in
that workbook (week52_high/low missing for some tickers - confirmed a data-
provider gap, not a formula bug; an implausible P/E on at least one ticker)
are reason enough to keep "never resave a file I don't fully understand" as
the standing rule, independent of whether resaving would actually be safe.

SANITY CHECKS FLAG, NEVER BLOCK. A suspect P/E, a stale as_of, or a missing
week52 range gets reported in the `flags` list on the output - the sweep
still runs. This mirrors the rest of the system's "no data is fine, never
invented, never silently substituted" rule.

CLAUDE-FOR-EXCEL FIX PROMPT (added 2026-08-11). When this run raises any
flags, they're also written as a short, ready-to-paste prompt to
data/cache/excel_import/claude_excel_prompt.txt - numbered fixes only,
nothing invented beyond what a flag already says, meant for the user's
Claude-for-Excel extension to act on directly inside the workbook. No
flags this run means no file (a stale one from a previous run is removed).
This script still never writes to the workbook itself - the prompt is
something the user hands to a *different* tool that can.

WHERE THE ACTUAL EXCEL FILE COMES FROM: this script takes a local path
(--xlsx). Getting the file itself out of Google Drive happens one level up,
in the Claude Code session: download via the Google Drive connector
(mcp__Google_Drive__download_file_content), base64-decode, save to
data/cache/excel_import/, then run this script against that path. Nothing
here talks to Drive directly.

Usage:
  python scripts/import_excel_holdings.py --xlsx /path/to/master-5.xlsx
  python scripts/import_excel_holdings.py --xlsx /path/to/master-5.xlsx --dry-run
"""
import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from config.settings import EXCEL_STALE_AFTER_DAYS, EXCEL_PE_SANITY_RANGE  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROFILES_DIR = os.path.join(ROOT, "data", "company_profiles")
PORTFOLIO_PATH = os.path.join(ROOT, "data", "portfolio.json")
TRANSACTIONS_CSV = os.path.join(ROOT, "data", "transactions.csv")
WATCHLIST_JSON = os.path.join(ROOT, "data", "cache", "watchlist.json")
IMPORT_SUMMARY_JSON = os.path.join(ROOT, "data", "cache", "excel_import", "latest-summary.json")
CLAUDE_EXCEL_PROMPT_PATH = os.path.join(ROOT, "data", "cache", "excel_import", "claude_excel_prompt.txt")

STOCK_DETAIL_MARKER = "STOCK DETAIL"
CORE_HOLDINGS_MARKER = "CORE HOLDINGS"
TRANSACTIONS_COLS = ["date", "type", "holdings_ticker", "name", "account", "quantity",
                     "price_per_unit", "amount_sek", "fee_sek", "cash_effect_sek",
                     "realized_pnl_sek", "source", "note"]
WATCHLIST_COLS = ["ticker", "name", "category", "price_sek", "pe_ratio", "market_cap",
                  "sector", "beta", "as_of", "notes"]


def _norm(v):
    return str(v).strip().lower() if v is not None else None


def find_section(ws, marker):
    """Row index whose row contains `marker` as a substring of any cell -
    that row is the section title; the header is assumed to be the next
    non-empty row. Returns (title_row, header_row) or (None, None)."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and marker in cell.value:
                return cell.row, cell.row + 1
    return None, None


def read_block(ws, header_row, required_cols):
    """Read rows below `header_row` as dicts keyed by header text (lowercased,
    matched against `required_cols`), stopping at the first row whose first
    matched required column is empty - the layout uses a blank spacer row
    between sections."""
    headers = [_norm(c.value) for c in ws[header_row]]
    col_idx = {}
    for want in required_cols:
        for i, h in enumerate(headers):
            if h == want:
                col_idx[want] = i
                break
    if not col_idx:
        return []
    key_col = col_idx[required_cols[0]]
    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        row_cells = ws[r]
        key_val = row_cells[key_col].value if key_col < len(row_cells) else None
        if key_val in (None, ""):
            break
        rec = {col: (row_cells[i].value if i < len(row_cells) else None)
               for col, i in col_idx.items()}
        rows.append(rec)
        r += 1
    return rows


def stale_days(as_of):
    if not isinstance(as_of, datetime):
        return None
    now = datetime.now(timezone.utc)
    ref = as_of if as_of.tzinfo else as_of.replace(tzinfo=timezone.utc)
    return (now - ref).days


def load_profile(ticker, name):
    path = os.path.join(PROFILES_DIR, f"{ticker}.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f), path
    return {
        "ticker": ticker, "name": name, "profile_last_updated": None,
        "static_profile": {"business_description": None, "sector": None,
                           "competitive_moat_notes": None, "management_notes": None,
                           "last_reviewed": None},
        "fundamentals_cache": {"source": None, "as_of_period": None,
                               "extracted_date": None, "next_report_expected": None,
                               "figures": {"ebit_margin_pct": None, "roic_pct": None,
                                          "roe_pct": None, "revenue_growth_pct": None,
                                          "fcf_sek": None, "net_debt_to_ebitda": None,
                                          "payout_ratio_pct": None}},
        "insider_activity_cache": {"source": None, "as_of_date": None, "notes": None},
        "review_history": [],
    }, path


def _structured(value, quality_state, as_of_str, age_days, source_tier=3,
                source="Excel Stocks data type (Drive, read-only)", calculation_method="direct from source, not computed"):
    """Builds one per-field entry in the company_profiles figures shape
    (see data/company_profiles/_SCHEMA.md, added 2026-08-09)."""
    return {"value": value, "source": source, "source_tier": source_tier,
            "as_of": as_of_str, "age_days": age_days, "quality_state": quality_state,
            "calculation_method": calculation_method}


def process_stock_detail(ws, today, flags, dry_run):
    """STOCK DETAIL block -> data/company_profiles/<TICKER>.json. Returns the
    set of tickers actually seen (for the data-gaps check against Holdings)."""
    title_row, header_row = find_section(ws, STOCK_DETAIL_MARKER)
    if title_row is None:
        flags.append("No 'STOCK DETAIL' block found in Holdings - nothing imported for fundamentals.")
        return set(), []

    cols = ["ticker", "name", "price_sek", "pe_ratio", "market_cap", "sector",
            "week52_high", "week52_low", "beta", "exchange", "currency",
            "headquarters", "employees", "as_of"]
    rows = read_block(ws, header_row, cols)
    seen, updated = set(), []
    pe_lo, pe_hi = EXCEL_PE_SANITY_RANGE

    for rec in rows:
        ticker, name = rec.get("ticker"), rec.get("name")
        if not ticker:
            continue
        seen.add(ticker)

        pe = rec.get("pe_ratio")
        if isinstance(pe, (int, float)) and not (pe_lo < pe < pe_hi):
            flags.append(f"{ticker}: P/E {pe:g} is outside the {pe_lo}-{pe_hi} sanity range - "
                         f"treat as suspect, verify in Excel before using it.")

        w_hi, w_lo = rec.get("week52_high"), rec.get("week52_low")
        if isinstance(w_hi, (int, float)) and isinstance(w_lo, (int, float)) and w_lo > w_hi:
            flags.append(f"{ticker}: week52_low ({w_lo}) > week52_high ({w_hi}) - inverted, verify in Excel.")
        elif not isinstance(w_hi, (int, float)) or not isinstance(w_lo, (int, float)):
            flags.append(f"{ticker}: no 52-week range in Excel (confirmed data-provider gap for some "
                         f"Nordic-primary tickers, not a formula bug) - treated as no data, not blocking.")

        age = stale_days(rec.get("as_of"))
        if age is not None and age > EXCEL_STALE_AFTER_DAYS:
            flags.append(f"{ticker}: Excel data is {age} days old (as_of {rec['as_of']}) - "
                         f"refresh the Stocks data type and re-save if you want current numbers.")

        profile, path = load_profile(ticker, name)
        as_of_str = rec["as_of"].strftime("%Y-%m-%d") if isinstance(rec.get("as_of"), datetime) else today
        if rec.get("sector"):
            profile["static_profile"]["sector"] = f"{rec['sector']} (Excel Stocks data type, {as_of_str})"

        figures = profile.setdefault("fundamentals_cache", {}).setdefault("figures", {})
        field_map = {"price_sek": "price", "pe_ratio": "pe_ratio", "market_cap": "market_cap",
                     "week52_high": "week52_high", "week52_low": "week52_low", "beta": "beta"}
        touched = []
        for excel_col, figure_key in field_map.items():
            v = rec.get(excel_col)
            if not isinstance(v, (int, float)):
                continue
            state = "OK"
            if figure_key == "pe_ratio" and not (pe_lo < v < pe_hi):
                state = "SUSPECT"
            elif age is not None and age > EXCEL_STALE_AFTER_DAYS:
                state = "STALE"
            figures[figure_key] = _structured(v, state, as_of_str, age)
            touched.append(figure_key)
        for extra_key, excel_col in (("exchange", "exchange"), ("currency", "currency"),
                                      ("headquarters", "headquarters"), ("employees", "employees")):
            if rec.get(excel_col) is not None:
                figures[extra_key] = _structured(rec[excel_col], "OK", as_of_str, age)

        if touched:
            profile["fundamentals_cache"]["source"] = "Excel Stocks data type (Drive, read-only)"
            profile["fundamentals_cache"]["extracted_date"] = as_of_str
        profile["profile_last_updated"] = today

        if not dry_run:
            os.makedirs(PROFILES_DIR, exist_ok=True)
            with open(path, "w") as f:
                json.dump(profile, f, indent=2)
                f.write("\n")
        updated.append((ticker, sorted(touched)))

    return seen, updated


def _match_key(ticker, account, name):
    """(ticker, account) uniquely identifies a holding EXCEPT for "TBD"
    tickers, which this system uses for multiple distinct funds in the same
    account (e.g. Avanza Auto 3 and Avanza Global are both "TBD"/avanza-isk).
    Fold name into the key in that case so they don't collide."""
    if ticker == "TBD":
        return (ticker, account, _norm(name))
    return (ticker, account, None)


def process_core_holdings(ws, pf, flags, dry_run):
    """CORE HOLDINGS block -> update data/portfolio.json quantity/cost basis.
    Every delta is reported, never silently applied."""
    title_row, header_row = find_section(ws, CORE_HOLDINGS_MARKER)
    if title_row is None:
        flags.append("No 'CORE HOLDINGS' block found in Holdings - portfolio.json not updated from Excel.")
        return [], set()

    cols = ["ticker", "name", "account", "quantity", "cost_basis_sek"]
    rows = read_block(ws, header_row, cols)
    by_key = {_match_key(h.get("ticker"), h.get("account_id"), h.get("name")): h
             for h in pf.get("holdings", [])}
    deltas = []
    seen = set()

    for rec in rows:
        ticker, account, name = rec.get("ticker"), rec.get("account"), rec.get("name")
        if not ticker or ticker == "TOTAL":
            continue
        seen.add(ticker)
        holding = by_key.get(_match_key(ticker, account, name))
        if holding is None:
            if ticker == "TBD":
                flags.append(f"Excel row \"{name}\" ({account}) has ticker TBD and no name match in "
                             f"portfolio.json - not updated, check the name matches exactly.")
            continue  # a ticker Excel knows about that isn't a tracked holding here - not an error

        qty, cost = rec.get("quantity"), rec.get("cost_basis_sek")
        old_qty = holding.get("quantity")
        if isinstance(qty, (int, float)) and qty != old_qty:
            deltas.append(f"{ticker} ({account}): quantity {old_qty!r} -> {qty!r} (from Excel)")
            if not dry_run:
                holding["quantity"] = qty
        if isinstance(cost, (int, float)) and cost != holding.get("cost_basis_total_sek") \
                and holding.get("cost_basis_per_unit") is None:
            deltas.append(f"{ticker} ({account}): cost_basis_total_sek "
                          f"{holding.get('cost_basis_total_sek')!r} -> {cost!r} (from Excel)")
            if not dry_run:
                holding["cost_basis_total_sek"] = cost

    return deltas, seen


def _find_header_row(ws, must_have):
    """Scan for the row whose cells contain every column name in `must_have`
    - the sheet may have a title/description row or two above the real
    header (Transactions does: title, then an append-only note, then headers)."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        vals = {_norm(c.value) for c in row}
        if must_have <= vals:
            return row[0].row
    return None


def process_transactions(ws, flags):
    """Transactions sheet -> append-only data/transactions.csv. Dedupes
    against what's already logged so re-running a sweep never duplicates."""
    header_row = _find_header_row(ws, {"date", "type", "holdings_ticker"})
    if header_row is None:
        flags.append("Transactions sheet header not recognized - nothing imported.")
        return []
    headers = [_norm(c.value) for c in ws[header_row]]
    col_idx = {}
    for want in TRANSACTIONS_COLS:
        for i, h in enumerate(headers):
            if h == want:
                col_idx[want] = i
                break

    def key_val(v):
        # Normalizes both sides of the dedup comparison the same way: a
        # blank Excel cell (None) and a blank CSV field (empty string after
        # DictReader) must produce the SAME key, or every re-run "discovers"
        # the same rows as new. str(None) == "None" is exactly the trap.
        return "" if v is None else str(v)

    existing = set()
    if os.path.exists(TRANSACTIONS_CSV):
        with open(TRANSACTIONS_CSV, newline="") as f:
            for row in csv.DictReader(f):
                existing.add(tuple(key_val(row.get(c)) for c in
                            ("date", "type", "holdings_ticker", "account", "quantity", "price_per_unit")))

    new_rows = []
    for row_cells in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        rec = {col: (row_cells[i].value if i < len(row_cells) else None) for col, i in col_idx.items()}
        if not rec.get("date") and not rec.get("type"):
            continue
        raw_date = rec.get("date")
        date_str = raw_date.strftime("%Y-%m-%d") if isinstance(raw_date, datetime) else key_val(raw_date)
        key = (date_str, key_val(rec.get("type")), key_val(rec.get("holdings_ticker")),
               key_val(rec.get("account")), key_val(rec.get("quantity")), key_val(rec.get("price_per_unit")))
        if key in existing:
            continue
        rec["date"] = date_str
        new_rows.append(rec)
        existing.add(key)

    return new_rows


def append_transactions(new_rows, dry_run):
    if not new_rows:
        return
    write_header = not os.path.exists(TRANSACTIONS_CSV)
    if dry_run:
        return
    os.makedirs(os.path.dirname(TRANSACTIONS_CSV), exist_ok=True)
    with open(TRANSACTIONS_CSV, "a", newline="") as f:
        w = csv.DictWriter(f, fieldnames=TRANSACTIONS_COLS)
        if write_header:
            w.writeheader()
        for rec in new_rows:
            w.writerow({k: rec.get(k, "") for k in TRANSACTIONS_COLS})


def process_watchlist(wb, flags, dry_run):
    if "Watchlist" not in wb.sheetnames:
        flags.append("No 'Watchlist' sheet in this workbook yet - data/universe.json stays in use "
                     "until it's added. See the Watchlist spec (ticker, name, category, price_sek, "
                     "pe_ratio, market_cap, sector, beta, as_of, notes).")
        return None
    ws = wb["Watchlist"]
    header_row = _find_header_row(ws, {"ticker"})
    if header_row is None:
        flags.append("Watchlist sheet found but has no 'ticker' column - not imported.")
        return None
    headers = [_norm(c.value) for c in ws[header_row]]
    col_idx = {w: headers.index(w) for w in WATCHLIST_COLS if w in headers}

    entries = []
    for row_cells in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        rec = {col: (row_cells[i].value if i < len(row_cells) else None) for col, i in col_idx.items()}
        if not rec.get("ticker"):
            continue
        if isinstance(rec.get("as_of"), datetime):
            rec["as_of"] = rec["as_of"].strftime("%Y-%m-%d")
        entries.append(rec)

    # "categories" groups tickers by the Watchlist's category column, in the
    # same {category: [tickers]} shape data/universe.json used - this is
    # what scripts/funnel/screen_candidates.py actually reads for a screen.
    # "entries" carries the full per-ticker record (including Excel's own
    # fundamentals and any per-row notes) for anything that wants more than
    # just the ticker list.
    categories = {}
    for rec in entries:
        cat = rec.get("category") or "uncategorized"
        categories.setdefault(cat, []).append(rec["ticker"])

    if not dry_run:
        os.makedirs(os.path.dirname(WATCHLIST_JSON), exist_ok=True)
        with open(WATCHLIST_JSON, "w") as f:
            json.dump({"source": "Watchlist tab, master-5.xlsx (Drive, read-only)",
                      "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                      "categories": categories,
                      "entries": entries}, f, indent=2)
    return entries


def build_claude_excel_prompt(flags, source_xlsx):
    """Turn this run's `flags` into a short, concrete prompt for the user's
    Claude-for-Excel extension to act on inside the live workbook. This is
    a format/data-quality check translated into fix instructions, not a new
    check of its own - every line here already traces to a flag this run
    raised. Written only when there's something to fix; no flags means no
    file, not an empty "everything's fine" prompt."""
    if not flags:
        return None
    lines = [
        "You're working inside my personal finance workbook "
        f"({os.path.basename(source_xlsx)}). Please make only the specific "
        "fixes below - don't change anything else in the sheet, and don't "
        "touch any linked Stocks data type cells beyond refreshing them "
        "where asked:",
        "",
    ]
    for i, fl in enumerate(flags, 1):
        lines.append(f"{i}. {fl}")
    lines.append("")
    lines.append(
        "After making a fix, leave the cell's normal formatting/data type "
        "in place - this workbook is read by a script that expects the "
        "existing column layout, so don't restructure a sheet to fix a row."
    )
    return "\n".join(lines) + "\n"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"{args.xlsx} not found.")

    wb = load_workbook(args.xlsx, data_only=True)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    flags = []

    if "Holdings" not in wb.sheetnames:
        raise SystemExit("No 'Holdings' sheet in this workbook.")
    ws_holdings = wb["Holdings"]
    fundamentals_seen, fundamentals_updated = process_stock_detail(ws_holdings, today, flags, args.dry_run)

    with open(PORTFOLIO_PATH) as f:
        pf = json.load(f)
    deltas, core_seen = process_core_holdings(ws_holdings, pf, flags, args.dry_run)
    if deltas and not args.dry_run:
        with open(PORTFOLIO_PATH, "w") as f:
            json.dump(pf, f, indent=2)
            f.write("\n")

    tx_new = []
    if "Transactions" in wb.sheetnames:
        tx_new = process_transactions(wb["Transactions"], flags)
        append_transactions(tx_new, args.dry_run)
    else:
        flags.append("No 'Transactions' sheet in this workbook - nothing imported.")

    watchlist_entries = process_watchlist(wb, flags, args.dry_run)

    # Data gaps: a tracked equity/certificate holding with nothing matching
    # in Excel's STOCK DETAIL block at all - the thing to go log next.
    tracked_tickers = {h.get("ticker") for h in pf.get("holdings", [])
                       if h.get("instrument_type") in ("stock", "certificate")
                       and h.get("ticker") not in (None, "TBD")}
    missing_from_excel = sorted(tracked_tickers - fundamentals_seen)
    if missing_from_excel:
        flags.append(f"Held but not in Excel's STOCK DETAIL block: {', '.join(missing_from_excel)} - "
                     f"add these tickers with the Stocks data type for next sweep.")

    # Persist a machine-readable summary - council reads THIS, not console
    # output, per the system's "every claim traces to a file this session"
    # rule. Written even on --dry-run (to a run marked as such) so a test
    # run doesn't leave a stale real summary lying around.
    summary = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "dry_run": args.dry_run,
        "source_xlsx": os.path.abspath(args.xlsx),
        "fundamentals_updated": [{"ticker": t, "fields": f} for t, f in fundamentals_updated],
        "portfolio_deltas": deltas,
        "transactions_appended": len(tx_new),
        "watchlist_entries": len(watchlist_entries) if watchlist_entries is not None else None,
        "flags": flags,
    }
    os.makedirs(os.path.dirname(IMPORT_SUMMARY_JSON), exist_ok=True)
    with open(IMPORT_SUMMARY_JSON, "w") as f:
        json.dump(summary, f, indent=2)

    prompt_text = build_claude_excel_prompt(flags, args.xlsx)
    if prompt_text and not args.dry_run:
        with open(CLAUDE_EXCEL_PROMPT_PATH, "w") as f:
            f.write(prompt_text)
    elif os.path.exists(CLAUDE_EXCEL_PROMPT_PATH) and not args.dry_run:
        # No flags this run - remove the stale prompt so an old fix list
        # doesn't linger after the thing it described got fixed.
        os.remove(CLAUDE_EXCEL_PROMPT_PATH)

    print(f"{'[dry-run] ' if args.dry_run else ''}Fundamentals: "
         f"{len(fundamentals_updated)} ticker(s) processed from Excel.")
    for ticker, fields in fundamentals_updated:
        print(f"  {ticker:<12s} {', '.join(fields) if fields else '(no numeric fields filled)'}")
    print(f"Portfolio.json deltas: {len(deltas)}")
    for d in deltas:
        print(f"  {d}")
    print(f"Transactions: {len(tx_new)} new row(s) appended to data/transactions.csv")
    print(f"Watchlist: {len(watchlist_entries) if watchlist_entries is not None else 'not present'}")
    print(f"\nFlags ({len(flags)}):")
    for fl in flags:
        print(f"  - {fl}")
    if prompt_text and not args.dry_run:
        print(f"\nClaude-for-Excel fix prompt written to {CLAUDE_EXCEL_PROMPT_PATH}")
    if args.dry_run:
        print("\n--dry-run: nothing written.")


if __name__ == "__main__":
    main()
