#!/usr/bin/env python3
"""
Bulk-import a flattened export of Excel's Stocks data type (or any other
manually-sourced field) into the Manual Data sheet.

WHY THIS EXISTS, AND WHY IT DOESN'T TOUCH THE STOCKS-LINKED FILE DIRECTLY:
Excel's Stocks data type / STOCKHISTORY function requires a genuine
Microsoft 365 subscription and a live, internet-connected Excel session —
Microsoft's cloud service resolves the data, not the file. It has no
presence in this headless pipeline. Worse: openpyxl (what this whole
project's sync layer is built on) has documented, real-world evidence of
silently corrupting complex Excel features it doesn't understand (macros,
named ranges, rich formatting — see github.com/anthropics/claude-code
issue #22044). Rich/Linked Data Types are exactly that kind of feature.
sync.py opens and re-saves master.xlsx on nearly every command, so putting
Stocks-linked cells IN master.xlsx risks silently breaking their live
refresh the next time any script touches the file.

The safe pattern: maintain Stocks-linked cells in a SEPARATE file, entirely
in real Excel, never touched by this project's scripts. When you want that
data in the system, convert the live cells to static values in Excel
(Copy -> Paste Special -> Values Only strips the rich-data-type magic,
leaving plain numbers/text) and export/save as CSV or a plain .xlsx — no
live formulas, no rich data types, nothing openpyxl needs to understand
beyond ordinary cells. This script reads THAT flattened file and bulk-loads
it into Manual Data via sync.py's append_rows() — same never-overwrite,
always-`_manual_overrides`-tagged mechanism every other manual entry uses.

Input file: CSV or .xlsx with these columns (extra columns ignored, matches
Manual Data's own schema exactly):
  ticker, field, value, currency, as_of, source, notes

Usage:
  python scripts/import_excel_stocks_data.py --file my_export.csv
  python scripts/import_excel_stocks_data.py --file my_export.xlsx --sheet Sheet1
"""
import argparse
import csv
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "sync"))
from schema import SHEETS  # noqa: E402

MANUAL_DATA_COLS = SHEETS["Manual Data"]  # ticker, field, value, currency, as_of, source, notes


def _read_csv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path, sheet_name):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)  # data_only: read cached VALUES, never a formula string
    ws = wb[sheet_name] if sheet_name else wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def _coerce_number(value):
    """Manual Data's `value` column should be numeric when the field is
    numeric — csv.DictReader and openpyxl string cells both hand back str;
    convert if it parses as a number, leave as text (e.g. a sector name)
    otherwise. Never guess beyond a literal parse."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        try:
            return float(value) if "." in value else int(value)
        except ValueError:
            return value
    return value


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True, help="CSV or .xlsx with ticker,field,value,... columns")
    p.add_argument("--sheet", help="(.xlsx input only) source sheet name, defaults to the active sheet")
    p.add_argument("--xlsx", default="master.xlsx", help="target master.xlsx (default: master.xlsx)")
    p.add_argument("--dry-run", action="store_true", help="print what would be imported, don't write anything")
    args = p.parse_args()

    if args.file.lower().endswith(".csv"):
        raw_rows = _read_csv(args.file)
    elif args.file.lower().endswith((".xlsx", ".xlsm")):
        raw_rows = _read_xlsx(args.file, args.sheet)
    else:
        sys.exit(f"Unsupported file type: {args.file} (expected .csv or .xlsx)")

    rows, skipped = [], []
    for r in raw_rows:
        ticker, field, value = r.get("ticker"), r.get("field"), r.get("value")
        if not ticker or not field or value in (None, ""):
            skipped.append(r)
            continue
        rows.append({
            "ticker": str(ticker).strip(),
            "field": str(field).strip(),
            "value": _coerce_number(value),
            "currency": r.get("currency") or None,
            "as_of": r.get("as_of") or None,
            "source": r.get("source") or f"bulk import from {os.path.basename(args.file)}",
            "notes": r.get("notes") or None,
        })

    print(f"Parsed {len(rows)} usable row(s) from {args.file}"
          f"{f', skipped {len(skipped)} (missing ticker/field/value)' if skipped else ''}.")
    for r in rows:
        print(f"  {r['ticker']:<14s} {r['field']:<20s} {r['value']}")

    if args.dry_run:
        print("\n--dry-run: nothing written. Re-run without it to import for real.")
        return
    if not rows:
        print("Nothing to import.")
        return

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + "/../data/sync")
    from sync import append_rows  # noqa: E402
    append_rows(args.xlsx, "Manual Data", rows)
    print("\nRun 'python run.py sync' then 'python run.py fetch' to apply these as "
          "fallback values (only where the automated fetch has no data) and see them "
          "flow into the next coverage report / lens read.")


if __name__ == "__main__":
    main()
