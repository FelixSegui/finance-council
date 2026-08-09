#!/usr/bin/env python3
"""
Fold the Fundamentals tab (in master.xlsx, user-maintained via Excel's
Stocks data type - see build_workbook.py's module docstring for why that
tab exists and its real limitations) into data/company_profiles/<TICKER>.json,
the cache the swedish-equity-review skill actually reads.

Reads the workbook's cached cell VALUES only (data_only=True) - this cannot
see or preserve a live Stocks data type link, only whatever Excel itself
last resolved and cached before the file was saved. That's fine: this
script exists to capture a periodic manual snapshot, not to run a live feed.

Field mapping: `sector` maps to the profile's static_profile.sector (a
direct match). Everything else (price, pe_ratio, market_cap,
dividend_yield_pct, week52_high, week52_low, beta) is written into
fundamentals_cache.figures under its OWN name, alongside (not instead of)
the structured figures (ebit_margin_pct, roic_pct, etc.) that
swedish-equity-review's schema defines - a P/E ratio is not an ROIC, and
this script never invents a conversion between unrelated metrics. Every
value written is tagged with its source and today's date.

Rows with nothing filled in beyond ticker/name are skipped - "the user
hasn't gotten to this one yet" is not the same as "confirmed no data".

Usage:
  python scripts/import_fundamentals_tab.py
  python scripts/import_fundamentals_tab.py --xlsx master.xlsx --dry-run
"""
import argparse
import json
import os
from datetime import datetime, timezone

from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROFILES_DIR = os.path.join(ROOT, "data", "company_profiles")

# Fundamentals-tab column -> where it lands. "sector" is a direct schema
# match; everything else goes into fundamentals_cache.figures verbatim,
# under its own name (see module docstring for why nothing gets remapped
# onto an unrelated existing figure key).
DIRECT_FIELDS = {"sector"}
FIGURE_FIELDS = ["price", "pe_ratio", "market_cap", "dividend_yield_pct",
                 "week52_high", "week52_low", "beta"]


def load_profile(ticker, name):
    path = os.path.join(PROFILES_DIR, f"{ticker}.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f), path
    return {
        "ticker": ticker, "name": name, "profile_last_updated": None,
        "static_profile": {"business_description": None, "sector": None,
                           "competitive_moat_notes": None, "management_notes": None,
                           "last_reviewed": None},
        "fundamentals_cache": {"source": None, "as_of_period": None,
                               "extracted_date": None, "next_report_expected": None,
                               "figures": {"ebit_margin_pct": None, "roic_pct": None,
                                          "roe_pct": None, "revenue_growth_pct": None,
                                          "fcf_sek": None, "net_debt_to_ebitda": None,
                                          "payout_ratio_pct": None}},
        "insider_activity_cache": {"source": None, "as_of_date": None, "notes": None},
        "review_history": [],
    }, path


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default=os.path.join(ROOT, "master.xlsx"))
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        raise SystemExit(f"{args.xlsx} not found — run scripts/build_workbook.py first.")
    wb = load_workbook(args.xlsx, data_only=True)
    if "Fundamentals" not in wb.sheetnames:
        raise SystemExit("No 'Fundamentals' sheet in this workbook.")
    ws = wb["Fundamentals"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    os.makedirs(PROFILES_DIR, exist_ok=True)
    updated = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(header, row))
        ticker, name = rec.get("ticker"), rec.get("name")
        if not ticker:
            continue
        filled = {k: v for k, v in rec.items() if k not in ("ticker", "name", "notes", "as_of")
                  and v not in (None, "")}
        if not filled:
            continue  # ticker seeded, nothing entered yet — not a data point

        profile, path = load_profile(ticker, name)
        as_of = rec.get("as_of") or today

        if "sector" in filled:
            profile["static_profile"]["sector"] = (
                f"{filled['sector']} (Excel Stocks data type, {as_of})")

        figures = profile.setdefault("fundamentals_cache", {}).setdefault("figures", {})
        touched_figures = []
        for f in FIGURE_FIELDS:
            if f in filled:
                # Structured per-field shape, see data/company_profiles/_SCHEMA.md
                # (added 2026-08-09) - source_tier 3 matches import_excel_holdings.py's
                # convention for the same underlying Excel Stocks data type source.
                figures[f] = {"value": filled[f], "source": "Excel Stocks data type (user's Microsoft 365 session)",
                              "source_tier": 3, "as_of": as_of, "age_days": None, "quality_state": "OK",
                              "calculation_method": "direct from source, not computed"}
                touched_figures.append(f)
        if touched_figures:
            profile["fundamentals_cache"]["source"] = "Excel Stocks data type (user's Microsoft 365 session)"
            profile["fundamentals_cache"]["extracted_date"] = as_of

        profile["profile_last_updated"] = today
        updated.append((ticker, sorted(filled.keys())))

        if not args.dry_run:
            with open(path, "w") as f:
                json.dump(profile, f, indent=2)

    if not updated:
        print("No filled-in Fundamentals rows found — nothing to import.")
        return
    print(f"{'Would update' if args.dry_run else 'Updated'} {len(updated)} profile(s):")
    for ticker, fields in updated:
        print(f"  {ticker:<14s} {', '.join(fields)}")
    if args.dry_run:
        print("\n--dry-run: nothing written.")


if __name__ == "__main__":
    main()
