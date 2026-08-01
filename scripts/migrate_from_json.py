#!/usr/bin/env python3
"""
ONE-TIME migration: reads the pre-migration portfolio.json,
investor_profile.json, thesis_candidates.json, universe.json from
archive/pre-migration-data/ (the frozen snapshot taken at migration time —
see SYSTEM.md) and populates master.xlsx.

This is intentionally a direct, hardcoded transcription script, not a generic
JSON-to-Excel framework — it ran once, by design (see SYSTEM.md's "smallest
system that works" principle). master.xlsx is the source of truth now; this
script is not part of any regular sweep, and is kept mainly so the migration
is re-runnable for verification (e.g. rebuilding master.xlsx from scratch
reproduces the same starting state).

Usage:
  python scripts/migrate_from_json.py                 # writes ./master.xlsx
  python scripts/migrate_from_json.py --out other.xlsx # for testing
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "sync"))
from workbook import build_workbook  # noqa: E402
from schema import SHEETS  # noqa: E402

from openpyxl import load_workbook

ARCHIVE_JSON_DIR = "archive/pre-migration-data"

# ---- risk-tier assignment (judgment call, not present as an explicit field
# in the current JSON — matches investor_profile.json's tier_definitions,
# which name these exact holdings by ticker) ----
TIER_BY_TICKER = {
    "TBD-Avanza Auto 3 (fund)": "secure",
    "TBD-Avanza Global (fund)": "secure",
    "SHB-A.ST": "medium",
    "INVE-A.ST": "medium",
    "COIN-XBT.ST": "high",
    "ethereum": "high",
}


def load(path):
    with open(path) as f:
        return json.load(f)


def cell(v):
    """openpyxl chokes on None being written as literal string 'None'; keep
    None as None (renders blank), everything else as-is."""
    return v


def populate_portfolio_sheet(ws, portfolio):
    accounts = {a["account_id"]: a for a in portfolio.get("accounts", [])}
    row = 2
    for h in portfolio.get("holdings", []):
        acc = accounts.get(h.get("account_id"), {})
        tier_key = f"{h.get('ticker')}-{h.get('name')}" if h.get("ticker") == "TBD" else h.get("ticker")
        risk_tier = TIER_BY_TICKER.get(tier_key, "cash" if h.get("exposure_class") == "cash" else "unassigned")
        vals = [
            h.get("ticker"), h.get("name"), h.get("account_id"),
            acc.get("institution"), acc.get("wrapper"),
            h.get("instrument_type"), h.get("exposure_class"), risk_tier,
            h.get("currency"), h.get("quantity"),
            h.get("cost_basis_per_unit"), h.get("cost_basis_total_sek"),
            h.get("annual_fee_pct"), h.get("date_acquired"),
            h.get("no_ticker_reason"),
            (h.get("thesis", "") or "")[:200] + ("…see Investment Thesis sheet" if h.get("thesis") else ""),
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=cell(v))
        row += 1
    return row - 2  # rows written


def populate_transactions_sheet(ws, portfolio):
    """Real historical trades, parsed from the scattered JSON where structured
    data exists (exit_plans, tax_breakdown_by_fund) plus the two most recent
    executed trades recorded in holding theses (Tundra sale, Avanza Global buy)."""
    rows = []

    hb = portfolio.get("exit_plans", {}).get("hb-af-exit", {}).get("actual_execution", {})
    if hb:
        d = hb.get("date")
        a50, a75 = hb.get("auto_50", {}), hb.get("auto_75", {})
        rows.append([d, "HB Auto 50 (fund)", "hb-main", "SELL", None, None, "SEK",
                     a50.get("sale_proceeds_sek"), a50.get("tax_at_30pct_sek"), "no",
                     "Genomsnittsmetoden; part of the HB fondkonto exit"])
        rows.append([d, "HB Auto 75 (fund)", "hb-main", "SELL", None, None, "SEK",
                     a75.get("sale_proceeds_sek"), a75.get("tax_at_30pct_sek"), "no",
                     "Genomsnittsmetoden; part of the HB fondkonto exit"])
    rows.append(["2026-07-20", "CASH_SEK", "hb-main→avanza-isk", "TRANSFER", None, None, "SEK",
                 127611.83, 0, "n/a", "HB fondkonto exit proceeds transferred to Avanza ISK"])

    for leg in portfolio.get("accounts", []):
        if leg.get("account_id") != "seb-fund":
            continue
        for fund in leg.get("tax_breakdown_by_fund", []):
            rows.append([fund.get("date"), fund.get("fund"), "seb-fund", "SELL", None, None, "SEK",
                         fund.get("sale_proceeds_sek"),
                         round(fund.get("kapitalvinst_sek", 0) * 0.30, 2), "no",
                         f"cost basis {fund.get('cost_basis_sek')} SEK, kapitalvinst {fund.get('kapitalvinst_sek')} SEK"])
    rows.append(["2026-07-20", "CASH_SEK", "seb-fund→avanza-isk", "TRANSFER", None, None, "SEK",
                 17382.43, 0, "n/a", "SEB fund proceeds transferred to Avanza ISK"])

    rows.append(["2026-07-29", "Tundra Sustainable Frontier Fund A SEK", "avanza-isk", "SELL", None, None, "SEK",
                 1534.86, 0, "yes",
                 "Sold in full — 2.6%/yr fee unjustifiable. Tax-free inside ISK. Cost basis 1,000 SEK."])
    rows.append(["2026-07-29", "Avanza Global (fund)", "avanza-isk", "BUY", None, None, "SEK",
                 119999, None, "yes",
                 "Secure-tier core purchase; ~59,050 SEK of this earmarked medium-pending for gradual migration"])

    for i, r in enumerate(rows, 2):
        for col, v in enumerate(r, 1):
            ws.cell(row=i, column=col, value=cell(v))
    return len(rows)


def populate_watchlist_sheet(ws, thesis_candidates):
    rows = thesis_candidates.get("candidates", [])
    for i, c in enumerate(rows, 2):
        vals = [c.get("ticker"), c.get("name"), c.get("date"), c.get("source"),
                f"risk_tag={c.get('risk_tag')}; policy_tailwind={c.get('policy_tailwind')}"]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=cell(v))
    return len(rows)


def populate_universe_sheet(ws, universe):
    manual_cats = universe.get("manual_categories", [c for c in universe.get("categories", {}) if c != "sp500"])
    meta = universe.get("metadata", {})
    row = 2
    n = 0
    for cat in manual_cats:
        for ticker in universe.get("categories", {}).get(cat, []):
            m = meta.get(ticker, {})
            vals = [ticker, cat, m.get("name"), m.get("currency"), None, m.get("source")]
            for col, v in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=cell(v))
            row += 1
            n += 1
    return n


def populate_thesis_sheet(ws, portfolio, thesis_candidates):
    row = 2
    n = 0
    for h in portfolio.get("holdings", []):
        if not h.get("thesis") or h.get("exposure_class") == "cash":
            continue
        ticker = h.get("ticker") if h.get("ticker") != "TBD" else h.get("name")
        vals = [ticker, portfolio.get("last_updated"), "see thesis text", None, None, h.get("thesis")]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=cell(v))
        row += 1
        n += 1
    for c in thesis_candidates.get("candidates", []):
        vals = [c.get("ticker"), c.get("date"), "candidate — not held", c.get("risk_tag"),
                c.get("policy_tailwind"), c.get("thesis")]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=cell(v))
        row += 1
        n += 1
    return n


def populate_pending_orders_sheet(ws, portfolio):
    items = portfolio.get("pending_executions", {}).get("items", [])
    for i, item in enumerate(items, 2):
        vals = [item.get("ticker"), item.get("name"), item.get("account_id"),
                item.get("committed_amount_sek"), item.get("committed_date"),
                item.get("status"), item.get("tier"), item.get("rationale")]
        for col, v in enumerate(vals, 1):
            ws.cell(row=i, column=col, value=cell(v))
    return len(items)


def populate_settings_sheet(ws, profile):
    risk = profile.get("risk", {})
    horizon = profile.get("horizon", {})
    buffer_ = profile.get("buffer", {})
    tier = profile.get("risk_tier_framework_proposed", {}).get("structure", {})
    routing = profile.get("risk_tier_framework_proposed", {}).get("contribution_routing", {})
    rows = [
        ("base_currency", "SEK", ""),
        ("max_drawdown_tolerance_pct", risk.get("max_drawdown_tolerance_pct"), risk.get("risk_capacity_notes", "")[:200]),
        ("primary_goal", horizon.get("primary_goal", "")[:200], ""),
        ("years_until_needed", horizon.get("years_until_needed"), "soft, not a hard deadline"),
        ("planned_monthly_contribution_sek", horizon.get("planned_monthly_contribution_sek"), "range, exact amount TBD"),
        ("emergency_fund_months", buffer_.get("emergency_fund_months"), buffer_.get("notes", "")),
        ("tier_secure_pct", tier.get("secure_non_volatile_pct"), "broad index funds"),
        ("tier_medium_pct", tier.get("medium_risk_pct"), "individual quality stocks"),
        ("tier_high_pct", tier.get("high_risk_pct"), "crypto, actively managed"),
        ("contribution_routing_default", routing.get("default_rule", "")[:200], ""),
        ("max_single_position_pct", profile.get("reference_targets", {}).get("max_single_position_pct"), ""),
        ("max_single_institution_pct", profile.get("reference_targets", {}).get("max_single_institution_pct"), ""),
        ("max_annual_fee_drag_pct", profile.get("reference_targets", {}).get("max_annual_fee_drag_pct"), ""),
    ]
    for i, r in enumerate(rows, 2):
        for col, v in enumerate(r, 1):
            ws.cell(row=i, column=col, value=cell(v))
    return len(rows)


def populate_notes_sheet(ws, portfolio):
    row = 2
    n = 0
    today = portfolio.get("last_updated")
    for q in portfolio.get("open_structural_questions", []):
        m = re.match(r"^(\d+)\.\s*(.*)$", q, re.DOTALL)
        qid, text = (m.group(1), m.group(2)) if m else (str(n + 1), q)
        ws.cell(row=row, column=1, value=cell(qid))
        ws.cell(row=row, column=2, value=cell(today))
        ws.cell(row=row, column=3, value=cell("open"))
        ws.cell(row=row, column=4, value=cell(text))
        row += 1
        n += 1
    for r in portfolio.get("resolved_structural_questions", []):
        m = re.match(r"^RESOLVED (\S+):\s*(.*)$", r, re.DOTALL)
        date, text = (m.group(1), m.group(2)) if m else (None, r)
        ws.cell(row=row, column=1, value=cell(n + 1))
        ws.cell(row=row, column=2, value=cell(date))
        ws.cell(row=row, column=3, value=cell("resolved"))
        ws.cell(row=row, column=4, value=cell(text))
        ws.cell(row=row, column=5, value=cell(date))
        row += 1
        n += 1
    return n


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="master.xlsx")
    args = p.parse_args()

    # The originals were archived once this migration ran for real (see
    # ARCHIVE_JSON_DIR below) — read from there so this script stays
    # re-runnable for verification without needing the pre-migration files
    # to still sit at their old data/ root paths.
    portfolio = load(f"{ARCHIVE_JSON_DIR}/portfolio.json")
    profile = load(f"{ARCHIVE_JSON_DIR}/investor_profile.json")
    thesis_candidates = load(f"{ARCHIVE_JSON_DIR}/thesis_candidates.json")
    universe = load(f"{ARCHIVE_JSON_DIR}/universe.json")

    build_workbook(args.out)
    wb = load_workbook(args.out)

    counts = {}
    counts["Portfolio"] = populate_portfolio_sheet(wb["Portfolio"], portfolio)
    counts["Transactions"] = populate_transactions_sheet(wb["Transactions"], portfolio)
    counts["Watchlist"] = populate_watchlist_sheet(wb["Watchlist"], thesis_candidates)
    counts["Universe"] = populate_universe_sheet(wb["Universe"], universe)
    counts["Investment Thesis"] = populate_thesis_sheet(wb["Investment Thesis"], portfolio, thesis_candidates)
    counts["Pending Orders"] = populate_pending_orders_sheet(wb["Pending Orders"], portfolio)
    counts["Settings"] = populate_settings_sheet(wb["Settings"], profile)
    counts["Notes"] = populate_notes_sheet(wb["Notes"], portfolio)

    wb.save(args.out)

    print(f"\nMigrated into {args.out}:")
    for sheet, n in counts.items():
        print(f"  {sheet:<20s} {n:>3d} rows")
    print(f"\nSource data read from {ARCHIVE_JSON_DIR}/ (the frozen pre-migration snapshot).")


if __name__ == "__main__":
    main()
