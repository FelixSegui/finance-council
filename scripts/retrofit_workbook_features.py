#!/usr/bin/env python3
"""
One-time (but safely re-runnable) retrofit for an ALREADY-POPULATED
master.xlsx: adds the _ValueHistory sheet (seeded from data/valuations.csv
if present), applies the Portfolio instrument_type conditional formatting,
and rebuilds the Dashboard sheet — all via the same functions
data/sync/workbook.py uses for a fresh build, so there is exactly one place
that defines what these features look like.

Never touches Zone-1 data (Portfolio rows, Transactions, etc.) — only adds
a new Zone-2 sheet if missing, re-applies formatting rules (idempotent,
safe to run repeatedly), and deletes+recreates the pure-formula Dashboard.

Usage:
  python scripts/retrofit_workbook_features.py --xlsx master.xlsx
"""
import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "sync"))
from workbook import (  # noqa: E402
    ensure_zone2_sheet, build_dashboard, _apply_portfolio_type_formatting,
)
from schema import SHEETS  # noqa: E402

from openpyxl import load_workbook

VALUATIONS_CSV = "data/valuations.csv"


def seed_value_history(ws):
    """If _ValueHistory is brand new (just the header row) and
    data/valuations.csv has history, copy it in once so existing manually-
    logged observations aren't lost. No-op if already seeded or if the CSV
    doesn't exist."""
    if ws.max_row > 1:
        return 0  # already has data — don't duplicate
    if not os.path.exists(VALUATIONS_CSV):
        return 0
    n = 0
    with open(VALUATIONS_CSV) as f:
        for row in csv.DictReader(f):
            n += 1
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=row["date"])
            ws.cell(row=r, column=2, value=float(row["total_value_sek"]))
            ws.cell(row=r, column=3, value=row.get("note", ""))
    return n


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default="master.xlsx")
    args = p.parse_args()

    wb = load_workbook(args.xlsx)

    vh = ensure_zone2_sheet(wb, "_ValueHistory")
    seeded = seed_value_history(vh)

    portfolio = wb["Portfolio"]
    portfolio.conditional_formatting._cf_rules.clear()  # idempotent: drop any prior rules first
    _apply_portfolio_type_formatting(portfolio, len(SHEETS["Portfolio"]))

    build_dashboard(wb)

    wb.save(args.xlsx)
    print(f"Retrofitted {args.xlsx}: _ValueHistory ({'seeded ' + str(seeded) + ' rows from ' + VALUATIONS_CSV if seeded else 'no seed needed'}), "
          f"Portfolio type-coloring, Dashboard rebuilt (BY TYPE + PERFORMANCE OVER TIME chart).")


if __name__ == "__main__":
    main()
